from django.http import HttpResponse
from django.contrib import admin, messages as mes
from import_export.admin import ImportExportMixin
from .models import Card, CardResource
from utils import format_card_number, format_phone_number, format_expire, format_balance
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import csv



@admin.register(Card)
class CardAdmin(ImportExportMixin, admin.ModelAdmin):
    resource_class = CardResource

    list_display = (
        'id',
        'display_card_number',
        'display_expire',
        'display_phone',
        'status',
        'display_balance',
    )
    search_fields = ('card_number', 'expire', 'phone', 'status', 'balance')
    list_filter = ('expire', 'phone', 'status', 'balance')

    @admin.display(description="Card number")
    def display_card_number(self, obj):
        return format_card_number(obj.card_number)

    @admin.display(description="Expire")
    def display_expire(self, obj):
        return format_expire(obj.expire)

    @admin.display(description="Phone")
    def display_phone(self, obj):
        return format_phone_number(obj.phone)

    @admin.display(description="Balance")
    def display_balance(self, obj):
        return format_balance(obj.balance)

    # === Actions ===
    actions = ["export_selected_xlsx", "export_selected_csv", "export_filtered_csv","send_fake_message"]

    # === Excel export (styled) ===
    def export_selected_xlsx(self, request, queryset):
        dataset = self.resource_class().export(queryset)

        wb = Workbook()
        ws = wb.active
        ws.title = "Cards"

        # Ustun sarlavhalari
        headers = dataset.headers
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ma’lumotlar
        for row_num, row_data in enumerate(dataset, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ustun kengligi auto
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 3

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="selected_cards.xlsx"'
        wb.save(response)
        return response

    export_selected_xlsx.short_description = "Export to Excel (selected rows)"

    # === CSV export ===
    def export_selected_csv(self, request, queryset):
        """Faqat tanlanganlarni export"""
        return self._export_to_csv(queryset, filename="selected_cards.csv")

    export_selected_csv.short_description = "Export to CSV (selected rows)"

    def export_filtered_csv(self, request, queryset):
        """Filterlangan butun querysetni export"""
        qs = self.get_queryset(request)  # bu yerda filterlangan queryset keladi
        return self._export_to_csv(qs, filename="filtered_cards.csv")

    export_filtered_csv.short_description = "Export to CSV (filtered data)"

    def _export_to_csv(self, queryset, filename):
        """Yagona CSV generator"""
        response = HttpResponse(content_type="text/csv")
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow(["ID", "Card Number", "Expire", "Phone", "Status", "Balance"])

        for obj in queryset:
            writer.writerow([
                obj.id,
                format_card_number(obj.card_number),
                format_expire(obj.expire),
                format_phone_number(obj.phone),
                obj.status,
                format_balance(obj.balance),
            ])

        return response

    def send_fake_message(self, request, queryset):
        from utils import send_otp

        messages = []
        for card in queryset:
            phone = str(card.phone).strip()
            if not phone.startswith("+"):
                phone = "+" + phone
            msg = (
                f"Sizning kartangiz {format_card_number(card.card_number)} aktiv va "
                f"foydalanishga {format_balance(card.balance)} UZS mavjud!"
            )
            messages.append((phone, msg))

        try:
            send_otp(7066090807,messages)
            self.message_user(request, f"✅ {len(messages)} ta xabar yuborildi!", mes.SUCCESS)
        except Exception as e:
            self.message_user(request, f"❌ Umumiy xatolik: {str(e)}", mes.ERROR)

    send_fake_message.short_description = "Send fake message to selected cards"

